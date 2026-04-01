import os
import threading
from dotenv import load_dotenv
load_dotenv()
import openpyxl

from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

from backend.database import init_db, get_db
from backend.firebase_setup import init_firebase
from backend.services import data_service
from backend.models.tournament import Tournament
from backend.routes import auth, matches, players, teams, scores, leaderboard, admin

app = FastAPI(title="Fantasy Cricket API")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Register API routes
app.include_router(auth.router)
app.include_router(matches.router)
app.include_router(players.router)
app.include_router(teams.router)
app.include_router(scores.router)
app.include_router(leaderboard.router)
app.include_router(admin.router)

# Tournament singleton
tournament = Tournament()
bootstrap_lock = threading.Lock()
bootstrap_started = False
bootstrap_error = None
bootstrap_ready = False


def seed_db_if_needed():
    db = get_db()
    row = db.execute("SELECT COUNT(*) as cnt FROM players").fetchone()
    count = row["cnt"] if isinstance(row, dict) else row[0]
    if count > 0:
        print(f"Database already has {count} players, skipping seed")
        return

    workbook_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "FantasyCricket.xlsx")
    if not os.path.exists(workbook_path):
        print(f"Seed file not found at {workbook_path}, skipping seed")
        return

    print("Seeding database from FantasyCricket.xlsx")
    wb = openpyxl.load_workbook(workbook_path, read_only=True)

    ws = wb["Players"]
    player_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        db.execute(
            """
            INSERT INTO players (id, name, team, role, aliases)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT (id) DO UPDATE SET
                name = ?,
                team = ?,
                role = ?,
                aliases = ?
            """,
            (int(row[0]), row[1], row[2], row[3], row[4] or "", row[1], row[2], row[3], row[4] or ""),
        )
        player_count += 1

    ws = wb["Matches"]
    match_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        date_str = row[1].strftime("%Y-%m-%d")
        time_str = row[2].strftime("%H:%M")
        db.execute(
            """
            INSERT INTO matches (id, team1, team2, match_date, match_time, status)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT (id) DO UPDATE SET
                team1 = ?,
                team2 = ?,
                match_date = ?,
                match_time = ?,
                status = ?
            """,
            (int(row[0]), row[3], row[4], date_str, time_str, "future", row[3], row[4], date_str, time_str, "future"),
        )
        match_count += 1

    db.commit()
    wb.close()
    print(f"Seeded: {player_count} players, {match_count} matches")


def bootstrap_app():
    global bootstrap_ready, bootstrap_error
    try:
        init_db()
        print("Database initialized")
        seed_db_if_needed()

        init_firebase()
        data_service.prime_static_cache()

        players_data = data_service.get_cached_data("players")
        matches_data = data_service.get_cached_data("matches")

        tournament.initialize(players_data, matches_data, [])
        tournament.start_scheduler()

        bootstrap_ready = True
        bootstrap_error = None
        print("Fantasy Cricket API started!")
    except Exception as exc:
        bootstrap_error = str(exc)
        print(f"Bootstrap error: {exc}")
        raise


def start_bootstrap_if_needed():
    global bootstrap_started
    with bootstrap_lock:
        if bootstrap_started:
            return
        bootstrap_started = True
        thread = threading.Thread(target=bootstrap_app, daemon=True)
        thread.start()


@app.on_event("startup")
def startup():
    admin.set_tournament(tournament)
    start_bootstrap_if_needed()


@app.get("/api/health")
def health():
    status = "ok" if bootstrap_ready else "starting"
    payload = {"status": status}
    if bootstrap_error:
        payload["bootstrap_error"] = bootstrap_error
    return payload


@app.get("/api/status")
def status():
    if not bootstrap_ready:
        return {
            "status": "starting",
            "scheduler": "booting",
            "bootstrap_error": bootstrap_error,
        }

    from backend.database import get_db
    db = get_db()
    try:
        players = db.execute("SELECT COUNT(*) as cnt FROM players").fetchone()
        matches = db.execute("SELECT COUNT(*) as cnt FROM matches").fetchone()
        users = db.execute("SELECT COUNT(*) as cnt FROM users").fetchone()
        return {
            "status": "ok",
            "players": players["cnt"] if isinstance(players, dict) else players[0],
            "matches": matches["cnt"] if isinstance(matches, dict) else matches[0],
            "users": users["cnt"] if isinstance(users, dict) else users[0],
            "scheduler": "running",
        }
    except Exception as e:
        return {"status": "error", "detail": str(e)}


# Serve React static files in production
STATIC_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "frontend", "dist")

if os.path.isdir(STATIC_DIR):
    # Serve static assets (js, css, images)
    app.mount("/assets", StaticFiles(directory=os.path.join(STATIC_DIR, "assets")), name="static-assets")

    # Catch-all: serve static files if they exist, otherwise index.html for React Router
    @app.get("/{full_path:path}")
    async def serve_spa(request: Request, full_path: str):
        if full_path.startswith("api/"):
            return {"detail": "Not found"}

        # Check if it's a real file in dist/
        file_path = os.path.join(STATIC_DIR, full_path)
        if full_path and os.path.isfile(file_path):
            return FileResponse(file_path)

        # Otherwise serve index.html for React Router
        return FileResponse(os.path.join(STATIC_DIR, "index.html"))
